"""
把「其他細項」欄的物品（如：水桶*1 鋼絲絨*1）拆成各自獨立欄位。
用法：python split_other_items.py 領用清單.xlsx
輸出：領用清單_拆欄.xlsx（同目錄）
"""
import sys
import re
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

def parse_items(text: str) -> dict:
    """解析「水桶×1 拖把桶*1 鋼絲絨×1」→ {'水桶': 1, '拖把桶': 1, ...}"""
    if not text:
        return {}
    text = str(text).strip()
    text = text.translate(str.maketrans({
        "×": "x",
        "＊": "*",
        "Ｘ": "x",
        "ｘ": "x",
        "０": "0",
        "１": "1",
        "２": "2",
        "３": "3",
        "４": "4",
        "５": "5",
        "６": "6",
        "７": "7",
        "８": "8",
        "９": "9",
    }))
    result = {}
    # 支援 x、* 當數量分隔；物品間用空白、全形空白、頓號、逗號或換行
    parts = re.split(r'[\s　\n、,，；;]+', text)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # 嘗試拆「名稱＋數量」
        m = re.match(r'^(.+?)[xX\*](\d+)(盒|個|包|組|支|條|張|本|瓶)?$', part)
        if m:
            name = m.group(1).strip()
            qty = int(m.group(2))
            result[name] = result.get(name, 0) + qty
        else:
            # 無數量標記→當作 1
            name = re.sub(r'[xX\*]\d+(盒|個|包|組|支|條|張|本|瓶)?$', '', part).strip()
            if name:
                result[name] = result.get(name, 0) + 1
    return result

def main(src: str):
    src_path = Path(src)
    if not src_path.exists():
        print(f"找不到檔案：{src}")
        sys.exit(1)

    wb = openpyxl.load_workbook(src_path)
    ws = wb.active

    # 找「其他細項」欄
    header_row = 1
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    try:
        other_col = headers.index("其他細項") + 1  # 1-based
    except ValueError:
        print("找不到「其他細項」欄，請確認標題名稱")
        sys.exit(1)

    # 收集所有物品名稱（統一順序）
    all_items: list[str] = []
    item_set: set[str] = set()
    row_items = []
    for r in range(header_row + 1, ws.max_row + 1):
        cell_val = ws.cell(r, other_col).value
        parsed = parse_items(cell_val)
        row_items.append(parsed)
        for name in parsed:
            if name not in item_set:
                item_set.add(name)
                all_items.append(name)

    if not all_items:
        print("「其他細項」欄沒有找到任何物品，請檢查格式（例如：水桶×1 拖把桶*1）")
        sys.exit(0)

    print(f"找到 {len(all_items)} 種物品：{all_items}")

    # 在「其他細項」欄之後插入新欄（原欄留著）
    insert_at = other_col + 1  # 插入位置（1-based → 插在「其他細項」後）

    # 插入欄（從右到左插，保持順序）
    for i, name in enumerate(reversed(all_items)):
        ws.insert_cols(insert_at)
        ws.cell(header_row, insert_at).value = name

    # 填入數量
    for r_idx, parsed in enumerate(row_items):
        row = header_row + 1 + r_idx
        for c_idx, name in enumerate(all_items):
            qty = parsed.get(name, 0)
            ws.cell(row, insert_at + c_idx).value = qty if qty else 0

    out_path = src_path.parent / (src_path.stem + "_拆欄.xlsx")
    wb.save(out_path)
    print(f"已儲存：{out_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法：python split_other_items.py 領用清單.xlsx")
    else:
        main(sys.argv[1])
